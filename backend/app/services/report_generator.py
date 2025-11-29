"""Premium report generation service for PDF and Excel exports"""
import io
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional

# Lazy imports for heavy dependencies - only import when needed
def _import_matplotlib():
    """Lazy import matplotlib to avoid startup errors if not installed"""
    try:
        import matplotlib
        matplotlib.use('Agg')  # Use non-interactive backend
        import matplotlib.pyplot as plt
        return plt
    except ImportError:
        raise ImportError("matplotlib is required for premium reports. Please install it: pip install matplotlib")

def _import_reportlab():
    """Lazy import reportlab"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
        from reportlab.lib.enums import TA_CENTER
        return {
            'colors': colors,
            'A4': A4,
            'getSampleStyleSheet': getSampleStyleSheet,
            'ParagraphStyle': ParagraphStyle,
            'inch': inch,
            'SimpleDocTemplate': SimpleDocTemplate,
            'Table': Table,
            'TableStyle': TableStyle,
            'Paragraph': Paragraph,
            'Spacer': Spacer,
            'Image': Image,
            'PageBreak': PageBreak,
            'TA_CENTER': TA_CENTER
        }
    except ImportError:
        raise ImportError("reportlab is required for premium reports. Please install it: pip install reportlab")

def _import_openpyxl():
    """Lazy import openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import PieChart, BarChart, LineChart, Reference
        from openpyxl.utils import get_column_letter
        return {
            'Workbook': Workbook,
            'Font': Font,
            'PatternFill': PatternFill,
            'Alignment': Alignment,
            'Border': Border,
            'Side': Side,
            'PieChart': PieChart,
            'BarChart': BarChart,
            'LineChart': LineChart,
            'Reference': Reference,
            'get_column_letter': get_column_letter
        }
    except ImportError:
        raise ImportError("openpyxl is required for premium reports. Please install it: pip install openpyxl")

def _import_numpy():
    """Lazy import numpy"""
    try:
        import numpy as np
        return np
    except ImportError:
        raise ImportError("numpy is required for premium reports. Please install it: pip install numpy")


class PremiumReportGenerator:
    """Generate premium reports with charts and detailed analytics"""
    
    def __init__(self, user_data: Dict[str, Any], analytics_data: Dict[str, Any], 
                 transactions_data: List[Dict[str, Any]], user_info: Dict[str, Any]):
        self.user_data = user_data
        self.analytics_data = analytics_data
        self.transactions_data = transactions_data
        self.user_info = user_info
        self.currency = user_info.get('default_currency', 'USD')
        
    def generate_pdf(self) -> bytes:
        """Generate PDF report with charts and detailed information"""
        # Import reportlab only when needed
        reportlab = _import_reportlab()
        SimpleDocTemplate = reportlab['SimpleDocTemplate']
        A4 = reportlab['A4']
        colors = reportlab['colors']
        getSampleStyleSheet = reportlab['getSampleStyleSheet']
        ParagraphStyle = reportlab['ParagraphStyle']
        inch = reportlab['inch']
        Table = reportlab['Table']
        TableStyle = reportlab['TableStyle']
        Paragraph = reportlab['Paragraph']
        Spacer = reportlab['Spacer']
        Image = reportlab['Image']
        PageBreak = reportlab['PageBreak']
        TA_CENTER = reportlab['TA_CENTER']
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#283593'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        elements.append(Paragraph("Финансовый отчет", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # User info and period
        user_name = self.user_info.get('first_name', '') or self.user_info.get('username', 'Пользователь')
        period_info = f"""
        <b>Пользователь:</b> {user_name}<br/>
        <b>Период:</b> {self.analytics_data.get('start_date', '')[:10]} - {self.analytics_data.get('end_date', '')[:10]}<br/>
        <b>Валюта:</b> {self.currency}
        """
        elements.append(Paragraph(period_info, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Executive Summary
        elements.append(Paragraph("Краткая сводка", heading_style))
        totals = self.analytics_data.get('totals', {})
        summary_data = [
            ['Показатель', 'Сумма'],
            ['Общий доход', f"{totals.get('income', 0):,.2f} {self.currency}"],
            ['Общие расходы', f"{totals.get('expense', 0):,.2f} {self.currency}"],
            ['Чистый поток', f"{totals.get('net', 0):,.2f} {self.currency}"],
            ['Количество транзакций', str(self.analytics_data.get('transaction_count', 0))]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3949ab')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Charts section
        elements.append(Paragraph("Визуализация данных", heading_style))
        
        # Generate and add charts
        charts = self._generate_charts()
        for chart_img in charts:
            elements.append(Image(chart_img, width=6*inch, height=4*inch))
            elements.append(Spacer(1, 0.2*inch))
        
        elements.append(PageBreak())
        
        # Top Categories
        elements.append(Paragraph("Топ категорий расходов", heading_style))
        expense_categories = self.analytics_data.get('top_expense_categories', [])
        if expense_categories:
            cat_data = [['Категория', 'Сумма', '% от общих расходов']]
            total_expense = totals.get('expense', 1)
            for cat in expense_categories[:10]:
                percentage = (cat.get('amount', 0) / total_expense * 100) if total_expense > 0 else 0
                cat_data.append([
                    f"{cat.get('icon', '')} {cat.get('name', '')}",
                    f"{cat.get('amount', 0):,.2f} {self.currency}",
                    f"{percentage:.1f}%"
                ])
            
            cat_table = Table(cat_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3949ab')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(cat_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Top Income Categories
        elements.append(Paragraph("Топ категорий доходов", heading_style))
        income_categories = self.analytics_data.get('top_income_categories', [])
        if income_categories:
            inc_data = [['Категория', 'Сумма', '% от общих доходов']]
            total_income = totals.get('income', 1)
            for cat in income_categories[:10]:
                percentage = (cat.get('amount', 0) / total_income * 100) if total_income > 0 else 0
                inc_data.append([
                    f"{cat.get('icon', '')} {cat.get('name', '')}",
                    f"{cat.get('amount', 0):,.2f} {self.currency}",
                    f"{percentage:.1f}%"
                ])
            
            inc_table = Table(inc_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
            inc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e7d32')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(inc_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Monthly Comparison
        elements.append(Paragraph("Месячное сравнение", heading_style))
        monthly_data = self.analytics_data.get('monthly_comparison', [])
        if monthly_data:
            month_data = [['Месяц', 'Доход', 'Расход', 'Чистый поток']]
            for month in monthly_data:
                month_data.append([
                    month.get('month', ''),
                    f"{month.get('income', 0):,.2f}",
                    f"{month.get('expense', 0):,.2f}",
                    f"{month.get('net', 0):,.2f}"
                ])
            
            month_table = Table(month_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            month_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5e35b1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lavender),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(month_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Goals Progress
        goals = self.analytics_data.get('goals', [])
        if goals:
            elements.append(PageBreak())
            elements.append(Paragraph("Прогресс по целям", heading_style))
            goals_data = [['Цель', 'Текущая сумма', 'Целевая сумма', 'Прогресс %', 'Осталось']]
            for goal in goals:
                goals_data.append([
                    goal.get('name', ''),
                    f"{goal.get('current_amount', 0):,.2f} {goal.get('currency', self.currency)}",
                    f"{goal.get('target_amount', 0):,.2f} {goal.get('currency', self.currency)}",
                    f"{goal.get('progress_percentage', 0):.1f}%",
                    f"{goal.get('remaining', 0):,.2f} {goal.get('currency', self.currency)}"
                ])
            
            goals_table = Table(goals_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch])
            goals_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c62828')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightcoral),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(goals_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Interesting Facts
        facts = self.analytics_data.get('facts', [])
        if facts:
            elements.append(Paragraph("Интересные факты", heading_style))
            facts_text = ""
            for fact in facts:
                facts_text += f"• {fact.get('icon', '')} {fact.get('text', '')}<br/>"
            elements.append(Paragraph(facts_text, styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
        
        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer_text = f"Отчет сгенерирован: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(footer_text, styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _generate_charts(self) -> List[io.BytesIO]:
        """Generate chart images for PDF"""
        # Import matplotlib only when needed
        plt = _import_matplotlib()
        np = _import_numpy()
        
        charts = []
        
        # Chart 1: Expenses by Category (Pie Chart)
        expense_categories = self.analytics_data.get('top_expense_categories', [])
        if expense_categories and len(expense_categories) > 0:
            fig, ax = plt.subplots(figsize=(8, 6))
            categories = [cat.get('name', '')[:15] for cat in expense_categories[:8]]
            amounts = [cat.get('amount', 0) for cat in expense_categories[:8]]
            colors_list = [cat.get('color', '#607D8B') for cat in expense_categories[:8]]
            
            ax.pie(amounts, labels=categories, autopct='%1.1f%%', startangle=90, colors=colors_list)
            ax.set_title('Распределение расходов по категориям', fontsize=14, fontweight='bold')
            plt.tight_layout()
            
            chart_buffer = io.BytesIO()
            plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
            chart_buffer.seek(0)
            charts.append(chart_buffer)
            plt.close()
        
        # Chart 2: Daily Flow (Line Chart)
        daily_flow = self.analytics_data.get('daily_flow', [])
        if daily_flow and len(daily_flow) > 0:
            fig, ax = plt.subplots(figsize=(10, 6))
            dates = [item.get('date', '') for item in daily_flow]
            income = [item.get('income', 0) for item in daily_flow]
            expense = [item.get('expense', 0) for item in daily_flow]
            
            ax.plot(dates, income, label='Доходы', color='#4CAF50', linewidth=2, marker='o')
            ax.plot(dates, expense, label='Расходы', color='#F44336', linewidth=2, marker='s')
            ax.set_xlabel('Дата', fontsize=12)
            ax.set_ylabel(f'Сумма ({self.currency})', fontsize=12)
            ax.set_title('Дневной денежный поток', fontsize=14, fontweight='bold')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            # Rotate x-axis labels if many dates
            if len(dates) > 10:
                plt.xticks(rotation=45, ha='right')
            
            plt.tight_layout()
            chart_buffer = io.BytesIO()
            plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
            chart_buffer.seek(0)
            charts.append(chart_buffer)
            plt.close()
        
        # Chart 3: Monthly Comparison (Bar Chart)
        monthly_data = self.analytics_data.get('monthly_comparison', [])
        if monthly_data and len(monthly_data) > 0:
            fig, ax = plt.subplots(figsize=(10, 6))
            months = [item.get('month_short', '') for item in monthly_data]
            income = [item.get('income', 0) for item in monthly_data]
            expense = [item.get('expense', 0) for item in monthly_data]
            
            x = np.arange(len(months))
            width = 0.35
            
            ax.bar(x - width/2, income, width, label='Доходы', color='#4CAF50')
            ax.bar(x + width/2, expense, width, label='Расходы', color='#F44336')
            ax.set_xlabel('Месяц', fontsize=12)
            ax.set_ylabel(f'Сумма ({self.currency})', fontsize=12)
            ax.set_title('Месячное сравнение доходов и расходов', fontsize=14, fontweight='bold')
            ax.set_xticks(x)
            ax.set_xticklabels(months)
            ax.legend()
            ax.grid(True, alpha=0.3, axis='y')
            
            plt.tight_layout()
            chart_buffer = io.BytesIO()
            plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
            chart_buffer.seek(0)
            charts.append(chart_buffer)
            plt.close()
        
        return charts
    
    def generate_excel(self) -> bytes:
        """Generate Excel report with charts and detailed data"""
        # Import openpyxl only when needed
        openpyxl = _import_openpyxl()
        Workbook = openpyxl['Workbook']
        Font = openpyxl['Font']
        PatternFill = openpyxl['PatternFill']
        Alignment = openpyxl['Alignment']
        Border = openpyxl['Border']
        Side = openpyxl['Side']
        PieChart = openpyxl['PieChart']
        BarChart = openpyxl['BarChart']
        LineChart = openpyxl['LineChart']
        Reference = openpyxl['Reference']
        get_column_letter = openpyxl['get_column_letter']
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Финансовый отчет"
        
        # Styles
        header_fill = PatternFill(start_color="3949ab", end_color="3949ab", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Финансовый отчет"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # User info
        row = 3
        user_name = self.user_info.get('first_name', '') or self.user_info.get('username', 'Пользователь')
        ws[f'A{row}'] = f"Пользователь: {user_name}"
        row += 1
        ws[f'A{row}'] = f"Период: {self.analytics_data.get('start_date', '')[:10]} - {self.analytics_data.get('end_date', '')[:10]}"
        row += 1
        ws[f'A{row}'] = f"Валюта: {self.currency}"
        row += 2
        
        # Executive Summary
        ws[f'A{row}'] = "Краткая сводка"
        ws[f'A{row}'].font = title_font
        row += 1
        
        summary_headers = ['Показатель', 'Сумма']
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        row += 1
        totals = self.analytics_data.get('totals', {})
        summary_data = [
            ['Общий доход', f"{totals.get('income', 0):,.2f} {self.currency}"],
            ['Общие расходы', f"{totals.get('expense', 0):,.2f} {self.currency}"],
            ['Чистый поток', f"{totals.get('net', 0):,.2f} {self.currency}"],
            ['Количество транзакций', str(self.analytics_data.get('transaction_count', 0))]
        ]
        
        for data_row in summary_data:
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
            row += 1
        
        row += 2
        
        # Top Expense Categories
        ws[f'A{row}'] = "Топ категорий расходов"
        ws[f'A{row}'].font = title_font
        row += 1
        
        expense_categories = self.analytics_data.get('top_expense_categories', [])
        if expense_categories:
            cat_headers = ['Категория', 'Сумма', '% от общих расходов']
            for col, header in enumerate(cat_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            total_expense = totals.get('expense', 1)
            for cat in expense_categories[:15]:
                percentage = (cat.get('amount', 0) / total_expense * 100) if total_expense > 0 else 0
                ws.cell(row=row, column=1).value = f"{cat.get('icon', '')} {cat.get('name', '')}"
                ws.cell(row=row, column=2).value = f"{cat.get('amount', 0):,.2f} {self.currency}"
                ws.cell(row=row, column=3).value = f"{percentage:.1f}%"
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = border
                row += 1
            
            # Pie Chart for expenses
            chart_start_row = row + 2
            pie_chart = PieChart()
            pie_chart.title = "Распределение расходов по категориям"
            data = Reference(ws, min_col=2, min_row=row-len(expense_categories), max_row=row-1)
            cats = Reference(ws, min_col=1, min_row=row-len(expense_categories), max_row=row-1)
            pie_chart.add_data(data, titles_from_data=False)
            pie_chart.set_categories(cats)
            pie_chart.width = 10
            pie_chart.height = 7
            ws.add_chart(pie_chart, f"E{chart_start_row}")
            row += 15
        
        row += 2
        
        # Top Income Categories
        ws[f'A{row}'] = "Топ категорий доходов"
        ws[f'A{row}'].font = title_font
        row += 1
        
        income_categories = self.analytics_data.get('top_income_categories', [])
        if income_categories:
            inc_headers = ['Категория', 'Сумма', '% от общих доходов']
            for col, header in enumerate(inc_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            total_income = totals.get('income', 1)
            for cat in income_categories[:15]:
                percentage = (cat.get('amount', 0) / total_income * 100) if total_income > 0 else 0
                ws.cell(row=row, column=1).value = f"{cat.get('icon', '')} {cat.get('name', '')}"
                ws.cell(row=row, column=2).value = f"{cat.get('amount', 0):,.2f} {self.currency}"
                ws.cell(row=row, column=3).value = f"{percentage:.1f}%"
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = border
                row += 1
            row += 2
        
        # Monthly Comparison
        ws[f'A{row}'] = "Месячное сравнение"
        ws[f'A{row}'].font = title_font
        row += 1
        
        monthly_data = self.analytics_data.get('monthly_comparison', [])
        if monthly_data:
            month_headers = ['Месяц', 'Доход', 'Расход', 'Чистый поток']
            for col, header in enumerate(month_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="5e35b1", end_color="5e35b1", fill_type="solid")
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            for month in monthly_data:
                ws.cell(row=row, column=1).value = month.get('month', '')
                ws.cell(row=row, column=2).value = month.get('income', 0)
                ws.cell(row=row, column=3).value = month.get('expense', 0)
                ws.cell(row=row, column=4).value = month.get('net', 0)
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border
                row += 1
            
            # Bar Chart for monthly comparison
            chart_start_row = row + 2
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.style = 10
            bar_chart.title = "Месячное сравнение доходов и расходов"
            bar_chart.y_axis.title = f"Сумма ({self.currency})"
            bar_chart.x_axis.title = "Месяц"
            
            data = Reference(ws, min_col=2, min_row=row-len(monthly_data)-1, max_col=3, max_row=row-1)
            cats = Reference(ws, min_col=1, min_row=row-len(monthly_data), max_row=row-1)
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)
            bar_chart.width = 15
            bar_chart.height = 7
            ws.add_chart(bar_chart, f"F{chart_start_row}")
            row += 15
        
        row += 2
        
        # Daily Flow Sheet
        ws2 = wb.create_sheet(title="Дневной поток")
        daily_flow = self.analytics_data.get('daily_flow', [])
        if daily_flow:
            row2 = 1
            headers = ['Дата', 'Доход', 'Расход', 'Чистый поток']
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=row2, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row2 += 1
            for day in daily_flow:
                ws2.cell(row=row2, column=1).value = day.get('date', '')
                ws2.cell(row=row2, column=2).value = day.get('income', 0)
                ws2.cell(row=row2, column=3).value = day.get('expense', 0)
                net = day.get('income', 0) - day.get('expense', 0)
                ws2.cell(row=row2, column=4).value = net
                for col in range(1, 5):
                    ws2.cell(row=row2, column=col).border = border
                row2 += 1
            
            # Line Chart for daily flow
            line_chart = LineChart()
            line_chart.title = "Дневной денежный поток"
            line_chart.y_axis.title = f"Сумма ({self.currency})"
            line_chart.x_axis.title = "Дата"
            
            data = Reference(ws2, min_col=2, min_row=1, max_col=3, max_row=row2-1)
            cats = Reference(ws2, min_col=1, min_row=2, max_row=row2-1)
            line_chart.add_data(data, titles_from_data=True)
            line_chart.set_categories(cats)
            line_chart.width = 20
            line_chart.height = 10
            ws2.add_chart(line_chart, "F2")
        
        # Goals Sheet
        goals = self.analytics_data.get('goals', [])
        if goals:
            ws3 = wb.create_sheet(title="Цели")
            row3 = 1
            goal_headers = ['Цель', 'Текущая сумма', 'Целевая сумма', 'Прогресс %', 'Осталось']
            for col, header in enumerate(goal_headers, 1):
                cell = ws3.cell(row=row3, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="c62828", end_color="c62828", fill_type="solid")
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row3 += 1
            for goal in goals:
                ws3.cell(row=row3, column=1).value = goal.get('name', '')
                ws3.cell(row=row3, column=2).value = goal.get('current_amount', 0)
                ws3.cell(row=row3, column=3).value = goal.get('target_amount', 0)
                ws3.cell(row=row3, column=4).value = goal.get('progress_percentage', 0)
                ws3.cell(row=row3, column=5).value = goal.get('remaining', 0)
                for col in range(1, 6):
                    ws3.cell(row=row3, column=col).border = border
                row3 += 1
        
        # Transactions Sheet
        if self.transactions_data:
            ws4 = wb.create_sheet(title="Транзакции")
            row4 = 1
            trans_headers = ['Дата', 'Тип', 'Сумма', 'Валюта', 'Категория', 'Описание', 'Счет']
            for col, header in enumerate(trans_headers, 1):
                cell = ws4.cell(row=row4, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row4 += 1
            for trans in self.transactions_data[:1000]:  # Limit to 1000 transactions
                ws4.cell(row=row4, column=1).value = trans.get('transaction_date', '')[:10] if trans.get('transaction_date') else ''
                ws4.cell(row=row4, column=2).value = trans.get('transaction_type', '')
                ws4.cell(row=row4, column=3).value = float(trans.get('amount', 0))
                ws4.cell(row=row4, column=4).value = trans.get('currency', self.currency)
                ws4.cell(row=row4, column=5).value = trans.get('category_name', '')
                ws4.cell(row=row4, column=6).value = trans.get('description', '')
                ws4.cell(row=row4, column=7).value = trans.get('account_name', '')
                for col in range(1, 8):
                    ws4.cell(row=row4, column=col).border = border
                row4 += 1
        
        # Adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

